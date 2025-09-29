# pdf_to_excel_format.py
import os
import camelot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# 🔹 Chemin complet vers Ghostscript
GS_PATH = r"C:\Program Files\gs\gs10.06.0\bin"
os.environ['PATH'] += f";{GS_PATH}"

def pdf_to_excel(pdf_path, excel_path):
    try:
        print(f"📂 Lecture du fichier PDF : {pdf_path} ...")
        
        # Camelot: flavor="stream" pour colonnes multiples
        tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream")

        if tables.n == 0:
            print("❌ Aucun tableau trouvé dans le PDF.")
            return

        # Exporter chaque tableau dans une feuille Excel
        tables.export("temp.xlsx", f="excel")

        # Amélioration de la mise en forme avec openpyxl
        wb = load_workbook("temp.xlsx")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Ajuster la largeur des colonnes
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
            
            # Mettre le header en gras et centré
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        wb.save(excel_path)
        print(f"✅ Conversion terminée ! Résultat final : {excel_path}")

    except Exception as e:
        print(f"⚠️ Erreur lors de la conversion : {e}")

if __name__ == "__main__":
    pdf_file = "JUIN_2025.pdf"
    excel_file = "resultat_format.xlsx"
    pdf_to_excel(pdf_file, excel_file)
