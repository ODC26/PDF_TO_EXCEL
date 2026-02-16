import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os
from datetime import datetime

# Chemin du fichier
fichier_entree = "precomptes_mupol.xlsx"
# Générer un nom de fichier avec horodatage
fichier_sortie = f"precomptes_mupol_traite_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

print(f"Lecture du fichier {fichier_entree}...")
# Lire le fichier Excel sans forcer les types (pour gérer les valeurs invalides)
df_original = pd.read_excel(fichier_entree)

# Ajouter une colonne avec le numéro de ligne d'origine (en tenant compte de l'en-tête Excel)
df_original['LIGNE_ORIGINE'] = range(2, len(df_original) + 2)

# Nettoyer et convertir les types de données
print("\nNettoyage et conversion des types de données...")

# NO_MATRICULE : convertir en numérique
df_original['NO_MATRICULE'] = pd.to_numeric(df_original['NO_MATRICULE'], errors='coerce')

# RECUP_NOM_AGENT : s'assurer que c'est du texte
df_original['RECUP_NOM_AGENT(NO_MATRICULE)'] = df_original['RECUP_NOM_AGENT(NO_MATRICULE)'].astype(str)

# MT_MENSUALITE : convertir en numérique
df_original['MT_MENSUALITE'] = pd.to_numeric(df_original['MT_MENSUALITE'], errors='coerce')

# NOMBRE_AYANTS_DROIT : Garder les valeurs telles quelles (ne pas convertir pour préserver "Montant invalide" etc.)
# Mais créer une version numérique pour les calculs si nécessaire
df_original['NOMBRE_AYANTS_DROIT_NUM'] = pd.to_numeric(df_original['NOMBRE_AYANTS_DROIT'], errors='coerce')

# Afficher les types de données après conversion
print(f"Nombre total de lignes : {len(df_original)}")
print(f"Colonnes : {list(df_original.columns)}")
print(f"\nTypes de données après conversion :")
print(df_original.dtypes)
print(f"\nLignes avec 'Montant invalide' ou texte dans NOMBRE_AYANTS_DROIT : {df_original['NOMBRE_AYANTS_DROIT'].apply(lambda x: isinstance(x, str) and not str(x).replace('.','').replace('-','').isdigit()).sum()}")

# Colonnes à vérifier pour les doublons complets
colonnes_toutes = ['NO_MATRICULE', 'RECUP_NOM_AGENT(NO_MATRICULE)', 'MT_MENSUALITE', 'NOMBRE_AYANTS_DROIT']

print("\nRecherche des doublons complets...")
# Ne pas exclure de lignes - garder TOUTES les lignes y compris les totaux
# Les NaN ne seront pas considérés comme doublons de toute façon
df_valides = df_original.copy()

print(f"Lignes analysées : {len(df_valides)}")

# Trouver les doublons complets (garder toutes les occurrences)
# dropna=False permet de ne pas considérer les NaN comme égaux
mask_doublons = df_valides.duplicated(subset=colonnes_toutes, keep=False)
df_doublons = df_valides[mask_doublons].copy()
df_doublons = df_doublons.sort_values(by=colonnes_toutes)

# Réorganiser les colonnes pour mettre LIGNE_ORIGINE en premier dans la feuille doublons
cols_doublons = ['LIGNE_ORIGINE'] + colonnes_toutes
df_doublons_export = df_doublons[cols_doublons]

print(f"Doublons trouvés : {len(df_doublons)} lignes")

# Créer le dataframe sans doublons (garder la première occurrence)
df_sans_doublons = df_valides.drop_duplicates(subset=colonnes_toutes, keep='first').copy()
# Sélectionner seulement les colonnes principales (sans NOMBRE_AYANTS_DROIT_NUM)
df_sans_doublons_export = df_sans_doublons[colonnes_toutes]
print(f"Lignes après suppression des doublons : {len(df_sans_doublons)}")

# Créer le fichier Excel avec les trois feuilles
print(f"\nCréation du nouveau fichier Excel avec 3 feuilles...")

with pd.ExcelWriter(fichier_sortie, engine='openpyxl') as writer:
    # Feuille 1 : Données originales (seulement les 4 colonnes principales)
    df_original[colonnes_toutes].to_excel(writer, sheet_name='DONNEES_INITIALES', index=False)
    print("✓ Feuille 'DONNEES_INITIALES' créée")
    
    # Feuille 2 : Doublons (avec la colonne LIGNE_ORIGINE)
    if len(df_doublons) > 0:
        df_doublons_export.to_excel(writer, sheet_name='DOUBLONS', index=False)
        print(f"✓ Feuille 'DOUBLONS' créée ({len(df_doublons)} lignes) avec colonne LIGNE_ORIGINE")
    else:
        # Créer une feuille vide avec les en-têtes
        cols_avec_ligne = ['LIGNE_ORIGINE'] + colonnes_toutes
        pd.DataFrame(columns=cols_avec_ligne).to_excel(writer, sheet_name='DOUBLONS', index=False)
        print("✓ Feuille 'DOUBLONS' créée (vide)")
    
    # Feuille 3 : Sans doublons
    df_sans_doublons_export.to_excel(writer, sheet_name='SANS_DOUBLONS', index=False)
    print(f"✓ Feuille 'SANS_DOUBLONS' créée ({len(df_sans_doublons)} lignes)")

# Ajuster la largeur des colonnes pour une meilleure lisibilité
print("\nAjustement de la largeur des colonnes...")
wb = load_workbook(fichier_sortie)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Ajuster la largeur des colonnes
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(fichier_sortie)
print("✓ Largeur des colonnes ajustée")

print("\n" + "="*80)
print("RÉSUMÉ")
print("="*80)
print(f"Fichier créé : {fichier_sortie}")
print(f"\nFeuille 'DONNEES_INITIALES' : {len(df_original)} lignes")
print(f"Feuille 'DOUBLONS' : {len(df_doublons)} lignes")
print(f"Feuille 'SANS_DOUBLONS' : {len(df_sans_doublons)} lignes")
print(f"\nDoublons supprimés : {len(df_valides) - len(df_sans_doublons)} lignes")
print(f"\nNote : Toutes les lignes sont conservées (y compris lignes de total et valeurs texte)")
print("="*80)
print("\n✅ Traitement terminé avec succès!")
