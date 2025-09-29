# pdf_to_excel_single_sheet.py
import os
import camelot
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# 🔹 Chemin complet vers Ghostscript
GS_PATH = r"C:\Program Files\gs\gs10.06.0\bin"
os.environ['PATH'] += f";{GS_PATH}"

def pdf_to_excel_single_sheet(pdf_path, excel_path):
    try:
        print(f"📂 Lecture du fichier PDF : {pdf_path} ...")

        # Camelot: flavor="stream" pour mieux détecter les colonnes
        tables = camelot.read_pdf(pdf_path, pages="all", flavor="stream")

        if tables.n == 0:
            print("❌ Aucun tableau trouvé dans le PDF.")
            return

        # Fusionner tous les tableaux dans un seul DataFrame
        all_dfs = []
        for table in tables:
            df = table.df
            # Retirer les colonnes vides et strip les headers
            df = df.dropna(axis=1, how='all')
            df.columns = [str(c).strip() for c in df.iloc[0]]
            df = df[1:]  # supprimer la ligne header répétée
            all_dfs.append(df)

        final_df = pd.concat(all_dfs, ignore_index=True)

        # Exporter dans Excel
        final_df.to_excel(excel_path, index=False, engine="openpyxl")

        # Amélioration mise en forme avec openpyxl
        wb = load_workbook(excel_path)
        ws = wb.active

        # Ajuster largeur des colonnes
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Mettre le header en gras et centré
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        wb.save(excel_path)
        print(f"✅ Conversion terminée ! Résultat final : {excel_path}")

    except Exception as e:
        print(f"⚠️ Erreur lors de la conversion : {e}")

if __name__ == "__main__":
    pdf_file = "JUIN_2025.pdf"
    excel_file = "resultat_unique.xlsx"
    pdf_to_excel_single_sheet(pdf_file, excel_file)

