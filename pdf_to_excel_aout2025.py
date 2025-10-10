import os
import sys
import argparse
import camelot
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# Path to Ghostscript (adjust if needed)
GS_PATH = r"C:\Program Files\gs\gs10.06.0\bin"
if os.path.isdir(GS_PATH):
    os.environ['PATH'] += f";{GS_PATH}"


def read_tables_with_fallback(pdf_path, pages="all"):
    """Try Camelot with stream then lattice as fallback. Returns camelot.TableList."""
    print("🔍 Tentative de lecture avec Camelot (stream)...")
    try:
        tables = camelot.read_pdf(pdf_path, pages=pages, flavor="stream")
        if tables.n > 0:
            print(f"✅ {tables.n} table(s) détectée(s) avec 'stream'.")
            return tables
    except Exception as e:
        print(f"⚠️ Erreur stream: {e}")

    print("🔁 Fallback vers 'lattice'...")
    try:
        tables = camelot.read_pdf(pdf_path, pages=pages, flavor="lattice")
        if tables.n > 0:
            print(f"✅ {tables.n} table(s) détectée(s) avec 'lattice'.")
            return tables
    except Exception as e:
        print(f"⚠️ Erreur lattice: {e}")

    return None


def sanitize_and_merge_tables(tables):
    """Convert a camelot TableList into a single cleaned DataFrame.

    Handles non-unique column names by appending suffixes, removes fully-empty columns,
    and drops repeated header rows if detected.
    """
    dfs = []
    for i, table in enumerate(tables):
        df = table.df.copy()
        # drop fully empty columns
        df = df.dropna(axis=1, how='all')
        if df.shape[0] == 0:
            continue

        # Detect header row candidates: often first row, but try to find a row with non-numeric vals
        header = df.iloc[0].astype(str).str.strip()
        # Make columns unique if not
        cols = list(header)
        if len(set(cols)) != len(cols):
            # append suffix to duplicates
            seen = {}
            new_cols = []
            for c in cols:
                key = c if c != '' else 'col'
                seen[key] = seen.get(key, 0) + 1
                if seen[key] > 1:
                    new_cols.append(f"{key}_{seen[key]}")
                else:
                    new_cols.append(key)
            cols = new_cols

        df.columns = cols
        # remove the header row from data
        df = df[1:].reset_index(drop=True)

        dfs.append(df)

    if not dfs:
        return pd.DataFrame()

    # concatenate, align columns
    combined = pd.concat(dfs, ignore_index=True, sort=False)

    # Reset columns: strip whitespace and fill empty names
    combined.columns = [str(c).strip() if str(c).strip() != '' else f'Column_{i}'
                        for i, c in enumerate(combined.columns, start=1)]

    # If duplicate column names remain, make them unique by adding suffix
    if combined.columns.duplicated().any():
        cols = []
        seen = {}
        for c in combined.columns:
            seen[c] = seen.get(c, 0) + 1
            if seen[c] > 1:
                cols.append(f"{c}_{seen[c]}")
            else:
                cols.append(c)
        combined.columns = cols

    return combined


def export_to_excel(df, excel_path):
    # If file exists, try to remove it first so openpyxl can write; if not possible, fallback to alternate name
    target_path = excel_path
    if os.path.exists(target_path):
        try:
            os.remove(target_path)
        except Exception:
            # Can't remove (maybe file open). We'll try to save to an alternate filename.
            base, ext = os.path.splitext(excel_path)
            from datetime import datetime
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            target_path = f"{base}_new_{ts}{ext}"
            print(f"⚠️ Ne peut pas écraser '{excel_path}'. Sauvegarde vers '{target_path}' à la place.")

    # Write DataFrame to excel
    try:
        df.to_excel(target_path, index=False, engine='openpyxl')
    except Exception as e:
        print(f"⚠️ Erreur lors de l'écriture initiale Excel: {e}")
        raise

    # Apply workbook formatting
    try:
        wb = load_workbook(target_path)
        ws = wb.active

        # Adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

        # Bold and center header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        wb.save(target_path)
    except PermissionError:
        # If still cannot save (file opened), inform user and leave the file written (or written to alternate name above)
        print(f"⚠️ Permission refusée lors de la sauvegarde du fichier Excel: '{target_path}'. Fermez le fichier s'il est ouvert et relancez.")
        raise
    except Exception as e:
        print(f"⚠️ Erreur lors de la mise en forme/sauvegarde Excel: {e}")
        raise

    # Return the actual saved path for caller use
    return target_path


def extract_trailing_text(pdf_path, pages="all"):
    """Extract textual content from PDF (all pages or specified pages) and return as list of lines."""
    lines = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if pages == 'all':
                page_iter = range(len(pdf.pages))
            else:
                # Convert pages like '1-3,5' into indices (0-based)
                page_iter = []
                parts = str(pages).split(',')
                for p in parts:
                    if '-' in p:
                        a, b = p.split('-')
                        page_iter.extend(range(int(a)-1, int(b)))
                    else:
                        page_iter.append(int(p)-1)

            for i in page_iter:
                if i < 0 or i >= len(pdf.pages):
                    continue
                page = pdf.pages[i]
                text = page.extract_text() or ''
                for l in text.splitlines():
                    if l.strip():
                        lines.append(l.strip())
    except Exception as e:
        print(f"⚠️ Erreur lors de l'extraction texte: {e}")

    return lines


def pdf_to_excel_robust(pdf_path, excel_path, pages="all", include_text=False):
    if not os.path.isfile(pdf_path):
        print(f"❌ Le fichier PDF spécifié n'existe pas: {pdf_path}")
        return

    print(f"📂 Lecture du fichier PDF : {pdf_path} ...")
    tables = read_tables_with_fallback(pdf_path, pages=pages)
    if not tables or tables.n == 0:
        print("❌ Aucun tableau trouvé dans le PDF après tentatives.")
        # But still optionally extract text
        if include_text:
            lines = extract_trailing_text(pdf_path, pages=pages)
            if lines:
                df = pd.DataFrame({'Extra_Text': lines})
                try:
                    saved = export_to_excel(df, excel_path)
                    print(f"✅ Export texte terminé : {saved}")
                except Exception as e:
                    print(f"⚠️ Erreur lors de l'export texte: {e}")
            else:
                print("❌ Aucun texte extrait non plus.")
        return

    df = sanitize_and_merge_tables(tables)
    if df.empty:
        print("❌ Aucun contenu tabulaire extrait après nettoyage.")
        # fallback to extracting text only
        if include_text:
            lines = extract_trailing_text(pdf_path, pages=pages)
            if lines:
                df = pd.DataFrame({'Extra_Text': lines})
                try:
                    saved = export_to_excel(df, excel_path)
                    print(f"✅ Export texte terminé : {saved}")
                except Exception as e:
                    print(f"⚠️ Erreur lors de l'export texte: {e}")
            else:
                print("❌ Aucun texte extrait non plus.")
        return

    # If requested, extract text lines and check for trailing lines not present in tables
    if include_text:
        text_lines = extract_trailing_text(pdf_path, pages=pages)
        if text_lines:
            # Heuristic: if last text line not present anywhere in df values, append as Extra_Text
            last_line = text_lines[-1]
            contains = df.apply(lambda col: col.astype(str).str.contains(last_line, na=False)).any().any()
            if not contains:
                # Append as a new row with only Extra_Text column
                df['Extra_Text'] = ''
                new_row = {c: '' for c in df.columns}
                new_row['Extra_Text'] = last_line
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True, sort=False)

    try:
        saved = export_to_excel(df, excel_path)
        print(f"✅ Conversion terminée ! Résultat final : {saved}")
    except Exception as e:
        print(f"⚠️ Erreur lors de l'export Excel : {e}")


def main():
    parser = argparse.ArgumentParser(description='Convert PDF tables to a single Excel sheet (robust).')
    parser.add_argument('pdf', help='Chemin vers le fichier PDF à convertir')
    parser.add_argument('-o', '--output', help='Chemin du fichier Excel de sortie', default='resultat_robuste.xlsx')
    parser.add_argument('-p', '--pages', help='Pages à analyser (ex: 1-3,5 or all)', default='all')
    parser.add_argument('--include-text', action='store_true', dest='include_text',
                        help="Inclure le texte non-tabulaire (extrait et ajouté en colonne 'Extra_Text' si absent des tables)")

    args = parser.parse_args()
    pdf_to_excel_robust(args.pdf, args.output, pages=args.pages, include_text=args.include_text)


if __name__ == '__main__':
    main()
