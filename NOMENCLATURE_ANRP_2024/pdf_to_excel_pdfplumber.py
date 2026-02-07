"""
Script de conversion PDF vers Excel utilisant pdfplumber
Extraction avancée avec détection automatique des tableaux
"""

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

def extract_tables_from_pdf(pdf_path):
    """
    Extrait tous les tableaux du PDF avec pdfplumber
    """
    all_tables = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"📄 Nombre de pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                print(f"📖 Traitement de la page {page_num}...")
                
                # Configuration de l'extraction des tableaux
                tables = page.extract_tables({
                    'vertical_strategy': 'lines',
                    'horizontal_strategy': 'lines',
                    'snap_tolerance': 3,
                    'join_tolerance': 3,
                    'edge_min_length': 3,
                    'min_words_vertical': 3,
                    'min_words_horizontal': 1,
                    'intersection_tolerance': 3,
                    'text_tolerance': 3,
                })
                
                if tables:
                    for table_num, table in enumerate(tables, start=1):
                        if table and len(table) > 1:  # Au moins une ligne d'en-tête + une ligne de données
                            all_tables.append({
                                'page': page_num,
                                'table_num': table_num,
                                'data': table
                            })
                            print(f"  ✓ Tableau {table_num} trouvé ({len(table)} lignes, {len(table[0]) if table else 0} colonnes)")
                else:
                    print(f"  ⚠ Aucun tableau trouvé sur la page {page_num}")
    
    except Exception as e:
        print(f"❌ Erreur lors de l'extraction: {e}")
        return []
    
    return all_tables

def clean_dataframe(df):
    """
    Nettoie le DataFrame en supprimant les lignes/colonnes vides
    """
    # Supprimer les lignes complètement vides
    df = df.dropna(how='all')
    
    # Supprimer les colonnes complètement vides
    df = df.dropna(axis=1, how='all')
    
    # Remplacer les None par des chaînes vides
    df = df.fillna('')
    
    # Nettoyer les espaces
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    
    return df

def format_excel(excel_path):
    """
    Applique une mise en forme professionnelle au fichier Excel
    """
    try:
        wb = load_workbook(excel_path)
        
        # Styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Formater la première ligne (en-tête)
            if ws.max_row > 0:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Formater les autres lignes
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.font = cell_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        
                        # Alterner les couleurs de fond
                        if row % 2 == 0:
                            cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            # Ajuster automatiquement la largeur des colonnes
            for col in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                
                for row in range(1, min(ws.max_row + 1, 100)):  # Limiter à 100 lignes pour la performance
                    cell = ws.cell(row=row, column=col)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                adjusted_width = min(max_length + 2, 50)  # Maximum 50 caractères
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Figer les volets (première ligne)
            ws.freeze_panes = ws['A2']
        
        wb.save(excel_path)
        print(f"✨ Mise en forme appliquée avec succès!")
        
    except Exception as e:
        print(f"⚠️ Erreur lors de la mise en forme: {e}")

def pdf_to_excel(pdf_path, excel_path, format_output=True):
    """
    Convertit un fichier PDF en Excel
    
    Args:
        pdf_path: Chemin du fichier PDF
        excel_path: Chemin du fichier Excel de sortie
        format_output: Appliquer une mise en forme (défaut: True)
    """
    if not os.path.exists(pdf_path):
        print(f"❌ Le fichier {pdf_path} n'existe pas!")
        return
    
    print(f"🚀 Démarrage de la conversion...")
    print(f"📂 Fichier source: {pdf_path}")
    print(f"📊 Fichier destination: {excel_path}")
    print("-" * 60)
    
    # Extraire les tableaux
    tables = extract_tables_from_pdf(pdf_path)
    
    if not tables:
        print("❌ Aucun tableau n'a été trouvé dans le PDF!")
        return
    
    print(f"\n✓ {len(tables)} tableau(x) extrait(s)")
    print("-" * 60)
    
    # Créer un fichier Excel avec plusieurs feuilles
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for idx, table_info in enumerate(tables):
                page = table_info['page']
                table_num = table_info['table_num']
                data = table_info['data']
                
                # Créer un DataFrame
                if data and len(data) > 0:
                    # Utiliser la première ligne comme en-tête
                    headers = data[0]
                    rows = data[1:]
                    
                    df = pd.DataFrame(rows, columns=headers)
                    df = clean_dataframe(df)
                    
                    # Nom de la feuille
                    sheet_name = f"Page{page}_Table{table_num}"
                    if len(sheet_name) > 31:  # Limite Excel
                        sheet_name = f"P{page}_T{table_num}"
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"✓ Feuille '{sheet_name}' créée ({len(df)} lignes)")
        
        print("-" * 60)
        print(f"✅ Conversion terminée!")
        
        # Appliquer la mise en forme
        if format_output:
            print("\n🎨 Application de la mise en forme...")
            format_excel(excel_path)
        
        print(f"\n📁 Fichier créé: {os.path.abspath(excel_path)}")
        
    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier Excel: {e}")

if __name__ == "__main__":
    # Configuration
    pdf_file = "NOMENCLATURE_NATIONALE_ANRP _ 2024.pdf"
    excel_file = "nomenclature_anrp_2024.xlsx"
    
    # Conversion
    pdf_to_excel(pdf_file, excel_file, format_output=True)
