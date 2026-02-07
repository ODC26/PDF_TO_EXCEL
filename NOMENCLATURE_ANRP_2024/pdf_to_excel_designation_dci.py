"""
Script d'extraction des colonnes Designation et DCI depuis un PDF
Extraction du texte brut et structuration en Excel
"""

import pdfplumber
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def extract_text_from_pdf(pdf_path):
    """
    Extrait tout le texte du PDF page par page
    """
    pages_data = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"📄 Nombre de pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text()
                if text:
                    pages_data.append({
                        'page': page_num,
                        'text': text
                    })
                    
                if page_num % 50 == 0:
                    print(f"📖 {page_num} pages traitées...")
    
    except Exception as e:
        print(f"❌ Erreur lors de l'extraction: {e}")
        return []
    
    print(f"✓ {len(pages_data)} pages extraites avec du texte")
    return pages_data

def parse_designation_dci(pages_data):
    """
    Parse le texte pour extraire les données Designation et DCI
    """
    all_records = []
    
    for page_info in pages_data:
        page_num = page_info['page']
        text = page_info['text']
        lines = text.split('\n')
        
        # Rechercher les lignes contenant "Designation" ou "DCI"
        for i, line in enumerate(lines):
            line_lower = line.lower()
            
            # Vérifier si la ligne contient des mots-clés de tableau
            if any(keyword in line_lower for keyword in ['designation', 'dci', 'dénomination']):
                # Extraire les données à partir de cette ligne
                # Les lignes suivantes contiennent probablement les données
                for j in range(i+1, min(i+50, len(lines))):  # Regarder jusqu'à 50 lignes après
                    data_line = lines[j].strip()
                    
                    if data_line and len(data_line) > 5:  # Ligne non vide avec contenu
                        # Essayer de diviser la ligne en colonnes
                        # Format typique: "designation ... dci ..."
                        parts = re.split(r'\s{2,}|\t', data_line)  # Split sur espaces multiples ou tabs
                        
                        if len(parts) >= 2:
                            all_records.append({
                                'Page': page_num,
                                'Designation': parts[0].strip(),
                                'DCI': parts[1].strip() if len(parts) > 1 else '',
                                'Autres': ' | '.join(parts[2:]) if len(parts) > 2 else ''
                            })
                
                break  # Passer à la page suivante après avoir trouvé le tableau
    
    return all_records

def extract_tables_with_text(pdf_path):
    """
    Extrait les tableaux en utilisant la méthode de détection de texte
    """
    all_data = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"📄 Nombre de pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                # Extraire le texte
                text = page.extract_text()
                
                if not text:
                    continue
                
                text_lower = text.lower()
                
                # Vérifier si la page contient les colonnes
                if 'designation' not in text_lower and 'dci' not in text_lower:
                    continue
                
                # Extraire ligne par ligne
                lines = text.split('\n')
                in_table = False
                designation_col = -1
                dci_col = -1
                header_positions = []
                
                for line_idx, line in enumerate(lines):
                    line_lower = line.lower()
                    
                    # Détecter l'en-tête du tableau
                    if 'designation' in line_lower or 'désignation' in line_lower:
                        in_table = True
                        
                        # Trouver les positions approximatives des colonnes
                        if 'designation' in line_lower:
                            designation_col = line_lower.find('designation')
                        elif 'désignation' in line_lower:
                            designation_col = line_lower.find('désignation')
                        
                        if 'dci' in line_lower:
                            dci_col = line_lower.find('dci')
                        
                        # Si on trouve les deux colonnes dans la même ligne
                        if designation_col >= 0 and dci_col >= 0:
                            print(f"📖 Page {page_num}: En-tête détecté (Designation pos {designation_col}, DCI pos {dci_col})")
                        continue
                    
                    # Extraire les données si on est dans un tableau
                    if in_table and line.strip():
                        # Arrêter si nouvelle section ou ligne trop courte
                        if any(keyword in line_lower for keyword in ['page ', 'chapitre', 'section', 'total', 'nomenclature']):
                            in_table = False
                            continue
                        
                        if len(line.strip()) < 5:
                            continue
                        
                        # Méthode 1: Utiliser les positions des colonnes détectées
                        if designation_col >= 0 and dci_col >= 0:
                            # Extraire selon les positions
                            designation = line[:dci_col].strip() if len(line) > designation_col else ''
                            dci = line[dci_col:].strip() if len(line) > dci_col else ''
                            
                            # Nettoyer et diviser si nécessaire
                            designation_parts = re.split(r'\s{3,}|\t', designation)
                            dci_parts = re.split(r'\s{3,}|\t', dci)
                            
                            designation_val = designation_parts[0] if designation_parts else ''
                            dci_val = dci_parts[0] if dci_parts else ''
                        else:
                            # Méthode 2: Split par espaces multiples
                            parts = re.split(r'\s{3,}|\t', line.strip())
                            
                            if len(parts) < 2:
                                # Essayer un autre split
                                parts = re.split(r'\s{2,}', line.strip())
                            
                            if len(parts) >= 2:
                                designation_val = parts[0].strip()
                                dci_val = parts[1].strip()
                            else:
                                continue
                        
                        # Ajouter si on a des données valides
                        if designation_val and len(designation_val) > 2 and not designation_val.lower().startswith('page'):
                            all_data.append({
                                'Page': page_num,
                                'Designation': designation_val,
                                'DCI': dci_val
                            })
                
                if page_num % 50 == 0:
                    print(f"  ... {page_num} pages traitées")
    
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return all_data

def format_excel(excel_path):
    """
    Applique une mise en forme au fichier Excel
    """
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater l'en-tête
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Formater les données
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Ajuster les largeurs de colonnes
        ws.column_dimensions['A'].width = 10  # Page
        ws.column_dimensions['B'].width = 60  # Designation
        ws.column_dimensions['C'].width = 60  # DCI
        
        # Figer les volets
        ws.freeze_panes = ws['A2']
        
        wb.save(excel_path)
        print("✨ Mise en forme appliquée!")
        
    except Exception as e:
        print(f"⚠️ Erreur mise en forme: {e}")

def pdf_to_excel(pdf_path, excel_path):
    """
    Convertit le PDF en Excel en extrayant les colonnes Designation et DCI
    """
    print("🚀 Extraction des colonnes Designation et DCI...")
    print(f"📂 Fichier: {pdf_path}")
    print("-" * 60)
    
    # Extraction des données
    data = extract_tables_with_text(pdf_path)
    
    if not data:
        print("❌ Aucune donnée Designation/DCI trouvée!")
        return
    
    print(f"\n✓ {len(data)} lignes extraites")
    print("-" * 60)
    
    # Créer le DataFrame
    df = pd.DataFrame(data)
    
    # Supprimer les lignes vides
    df = df[(df['Designation'].str.len() > 0) | (df['DCI'].str.len() > 0)]
    
    # Sauvegarder en Excel
    df.to_excel(excel_path, index=False, engine='openpyxl')
    print(f"✅ Fichier créé: {excel_path}")
    
    # Mise en forme
    print("\n🎨 Application de la mise en forme...")
    format_excel(excel_path)
    
    print(f"\n📊 Résumé:")
    print(f"  - Total de lignes: {len(df)}")
    print(f"  - Pages avec données: {df['Page'].nunique()}")
    print(f"  - Designation non vides: {df['Designation'].astype(bool).sum()}")
    print(f"  - DCI non vides: {df['DCI'].astype(bool).sum()}")

if __name__ == "__main__":
    pdf_file = "NOMENCLATURE_NATIONALE_ANRP _ 2024.pdf"
    excel_file = "nomenclature_designation_dci.xlsx"
    
    pdf_to_excel(pdf_file, excel_file)
