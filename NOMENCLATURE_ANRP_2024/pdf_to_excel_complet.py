"""
Script d'extraction complète du tableau avec toutes les colonnes
Gère les lignes multiples pour une même Designation
"""

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

def extract_full_table(pdf_path):
    """
    Extrait le tableau complet avec des colonnes fixées par les positions de l'en-tête
    """
    all_data = []
    column_positions = None
    column_names = []
    column_edges = None
    focus_columns = ['N°', 'Designation', 'DCI']
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"📄 Nombre de pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                # Détecter l'en-tête et les positions des colonnes
                if column_positions is None:
                    words = page.extract_words(x_tolerance=2, y_tolerance=2)
                    if words:
                        lines_dict = {}
                        for word in words:
                            y_pos = round(word['top'])
                            lines_dict.setdefault(y_pos, []).append(word)
                        
                        for y_pos, line_words in sorted(lines_dict.items()):
                            line_text = ' '.join([w['text'] for w in line_words]).lower()
                            if 'designation' in line_text or 'désignation' in line_text:
                                sorted_words = sorted(line_words, key=lambda w: w['x0'])
                                page_positions = {}
                                for word in sorted_words:
                                    w = word['text'].strip().lower()
                                    if w in ['n°', 'n', 'no', 'num'] or 'n°' in w:
                                        page_positions['N°'] = word['x0']
                                    elif 'designation' in w or 'désignation' in w:
                                        page_positions['Designation'] = word['x0']
                                    elif 'dci' in w:
                                        page_positions['DCI'] = word['x0']
                                    elif 'dosage' in w:
                                        page_positions['Dosage'] = word['x0']
                                    elif "admin" in w or "administr" in w:
                                        page_positions["Administration"] = word['x0']
                                    elif 'fabricant' in w:
                                        page_positions['Fabricant'] = word['x0']
                                    elif 'pght' in w:
                                        page_positions['PGHT'] = word['x0']
                                    elif 'cfa' in w:
                                        page_positions['CFA'] = word['x0']
                                    elif 'code' in w:
                                        page_positions['Code'] = word['x0']
                                    elif 'amm' in w:
                                        page_positions['AMM'] = word['x0']
                                    elif 'expiration' in w or "d'exp" in w:
                                        page_positions['Expiration'] = word['x0']
                                
                                if page_positions:
                                    column_positions = dict(sorted(page_positions.items(), key=lambda x: x[1]))
                                    column_names = list(column_positions.keys())
                                    # Calculer les bords de colonnes (milieux entre x0)
                                    x_positions = list(column_positions.values())
                                    edges = [max(0, x_positions[0] - 5)]
                                    for i in range(len(x_positions) - 1):
                                        edges.append((x_positions[i] + x_positions[i + 1]) / 2)
                                    edges.append(page.width - 2)
                                    column_edges = edges
                                    print(f"📋 Colonnes détectées: {column_names}")
                                    print(f"   Positions X: {column_positions}")
                                break
                
                if not column_edges:
                    continue
                
                # Extraction manuelle par lignes, avec colonnes explicites
                words = page.extract_words(x_tolerance=2, y_tolerance=2)
                if not words:
                    continue
                
                # Grouper par ligne
                lines_dict = {}
                for word in words:
                    y_pos = round(word['top'])
                    lines_dict.setdefault(y_pos, []).append(word)
                
                # Trouver la ligne d'en-tête sur cette page
                header_y = None
                for y_pos, line_words in sorted(lines_dict.items()):
                    line_text = ' '.join([w['text'] for w in line_words]).lower()
                    if 'designation' in line_text and 'dci' in line_text:
                        header_y = y_pos
                        break
                if header_y is None:
                    continue
                
                # Extraire le tableau avec colonnes explicites
                if 'N°' not in column_names or 'Designation' not in column_names or 'DCI' not in column_names:
                    continue
                
                n_idx = column_names.index('N°')
                d_idx = column_names.index('Designation')
                dci_idx = column_names.index('DCI')
                
                table_settings = {
                    'vertical_strategy': 'explicit',
                    'horizontal_strategy': 'text',
                    'explicit_vertical_lines': column_edges,
                    'snap_tolerance': 3,
                    'join_tolerance': 3,
                    'min_words_vertical': 1,
                    'min_words_horizontal': 1,
                }
                
                table = page.extract_table(table_settings)
                if not table:
                    continue
                
                # Trouver la ligne d'en-tête dans la table
                header_row_idx = None
                for idx, row in enumerate(table):
                    row_text = ' '.join([str(c).lower() if c else '' for c in row])
                    if ('design' in row_text and 'dci' in row_text) or ('designation' in row_text and 'dci' in row_text):
                        header_row_idx = idx
                        break
                if header_row_idx is None:
                    continue
                
                current_row = None
                for row in table[header_row_idx + 1:]:
                    if not row or not any(cell for cell in row):
                        continue
                    
                    row_text = ' '.join([str(c).lower() if c else '' for c in row])
                    if any(k in row_text for k in ['dosage', 'administrati', 'fabricant', 'classe', "d'amm", 'expiration']) and not re.search(r'\d', row_text):
                        continue
                    
                    # Normaliser la longueur de ligne
                    if len(row) < len(column_names):
                        row = row + [''] * (len(column_names) - len(row))
                    if len(row) > len(column_names):
                        row = row[:len(column_names)]
                    
                    numero = str(row[n_idx]).replace('\n', ' ').strip() if row[n_idx] else ''
                    designation = str(row[d_idx]).replace('\n', ' ').strip() if row[d_idx] else ''
                    dci = str(row[dci_idx]).replace('\n', ' ').strip() if row[dci_idx] else ''
                    
                    row_data = {'Page': page_num, 'N°': numero, 'Designation': designation, 'DCI': dci}
                    
                    if not any(row_data[col] for col in focus_columns):
                        continue
                    
                    # Lignes de continuation
                    if not numero and current_row:
                        if designation:
                            current_row['Designation'] = f"{current_row['Designation']} | {designation}" if current_row['Designation'] else designation
                        if dci:
                            current_row['DCI'] = f"{current_row['DCI']} | {dci}" if current_row['DCI'] else dci
                    else:
                        if current_row:
                            all_data.append(current_row)
                        current_row = row_data
                
                if current_row:
                    all_data.append(current_row)
                
                if page_num % 50 == 0:
                    print(f"  ... {page_num} pages traitées ({len(all_data)} lignes)")
    
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return all_data

def normalize_columns(data):
    """
    Normalise les noms de colonnes (pas besoin car déjà fait)
    """
    if not data:
        return []
    
    print(f"\n📋 Colonnes dans les données: {list(data[0].keys()) if data else []}")
    return data

def clean_data(data):
    """
    Nettoie les données
    """
    cleaned = []
    
    for row in data:
        # Filtrer les lignes complètement vides
        if not any(str(v).strip() for k, v in row.items() if k != 'Page'):
            continue
        
        # Filtrer les en-têtes répétés / pieds de page
        designation = str(row.get('Designation', '')).lower()
        dci = str(row.get('DCI', '')).lower()
        if 'designation' in designation or 'nomenclature' in designation:
            continue
        if 'page' in designation and 'de' in designation:
            continue
        if 'page' in dci and 'de' in dci:
            continue
        
        cleaned.append(row)
    
    return cleaned

def merge_continuations(rows):
    """
    Fusionne les lignes sans N° avec la ligne principale la plus proche.
    """
    merged = []
    pending = []
    
    for row in rows:
        numero = str(row.get('N°', '')).strip()
        if numero:
            if pending:
                for p in pending:
                    for col in ['Designation', 'DCI']:
                        if p.get(col):
                            if row.get(col):
                                row[col] = f"{p[col]} | {row[col]}"
                            else:
                                row[col] = p[col]
                pending = []
            merged.append(row)
        else:
            if merged:
                for col in ['Designation', 'DCI']:
                    if row.get(col):
                        if merged[-1].get(col):
                            merged[-1][col] = f"{merged[-1][col]} | {row[col]}"
                        else:
                            merged[-1][col] = row[col]
            else:
                pending.append(row)
    
    if pending and merged:
        first = merged[0]
        for p in pending:
            for col in ['Designation', 'DCI']:
                if p.get(col):
                    if first.get(col):
                        first[col] = f"{p[col]} | {first[col]}"
                    else:
                        first[col] = p[col]
    
    return merged

def merge_by_numero(rows):
    """
    Regroupe les lignes par N° pour éviter les doublons.
    """
    grouped = {}
    order = []
    for row in rows:
        numero = str(row.get('N°', '')).strip()
        if not numero:
            continue
        if numero not in grouped:
            grouped[numero] = row.copy()
            order.append(numero)
        else:
            for col in ['Designation', 'DCI']:
                if row.get(col):
                    if grouped[numero].get(col):
                        grouped[numero][col] = f"{grouped[numero][col]} | {row[col]}"
                    else:
                        grouped[numero][col] = row[col]
    return [grouped[n] for n in order]

def format_excel(excel_path):
    """
    Applique une mise en forme au fichier Excel
    """
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater l'en-tête
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Formater les données
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Ajuster les largeurs
        for col_idx, col in enumerate(ws.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[col_letter].width = max(adjusted_width, 10)
        
        # Figer les volets
        ws.freeze_panes = ws['A2']
        
        wb.save(excel_path)
        print("✨ Mise en forme appliquée!")
        
    except Exception as e:
        print(f"⚠️ Erreur mise en forme: {e}")

def pdf_to_excel(pdf_path, excel_path):
    """
    Convertit le PDF en Excel avec toutes les colonnes
    """
    print("🚀 Extraction complète du tableau...")
    print(f"📂 Fichier: {pdf_path}")
    print("=" * 60)
    
    # Extraction
    data = extract_full_table(pdf_path)
    
    if not data:
        print("❌ Aucune donnée trouvée!")
        return
    
    print(f"\n✓ {len(data)} lignes extraites")
    
    # Normalisation
    data = normalize_columns(data)
    
    # Fusionner les lignes sans N°
    data = merge_continuations(data)
    # Regrouper par N°
    data = merge_by_numero(data)
    
    # Nettoyage
    data = clean_data(data)
    print(f"✓ {len(data)} lignes après nettoyage")
    print("=" * 60)
    
    # Créer le DataFrame
    df = pd.DataFrame(data)
    
    # Réorganiser les colonnes dans un ordre logique
    desired_order = ['Page', 'N°', 'Designation', 'DCI']
    actual_columns = [col for col in desired_order if col in df.columns]
    other_columns = [col for col in df.columns if col not in desired_order]
    final_columns = actual_columns + other_columns
    
    df = df[final_columns]
    
    # Sauvegarder
    df.to_excel(excel_path, index=False, engine='openpyxl')
    print(f"✅ Fichier créé: {excel_path}")
    
    # Mise en forme
    print("\n🎨 Application de la mise en forme...")
    format_excel(excel_path)
    
    print(f"\n📊 Résumé:")
    print(f"  - Total de lignes: {len(df)}")
    print(f"  - Colonnes: {', '.join(df.columns)}")
    print(f"  - Pages avec données: {df['Page'].nunique()}")
    
    # Afficher un aperçu
    print(f"\n👀 Aperçu des 5 premières lignes:")
    print(df.head(5).to_string())

if __name__ == "__main__":
    pdf_file = "NOMENCLATURE_NATIONALE_ANRP _ 2024.pdf"
    excel_file = "nomenclature_complete.xlsx"
    
    pdf_to_excel(pdf_file, excel_file)
