"""
Script d'extraction avancée des colonnes Designation et DCI
Utilise la détection de colonnes visuelles de pdfplumber
"""

import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

def extract_table_data(pdf_path):
    """
    Extrait les données en utilisant la détection de tableaux améliorée
    """
    all_data = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"📄 Nombre de pages: {len(pdf.pages)}")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                # Configuration pour une meilleure extraction
                table_settings = {
                    'vertical_strategy': 'lines_strict',
                    'horizontal_strategy': 'lines_strict',
                    'explicit_vertical_lines': [],
                    'explicit_horizontal_lines': [],
                    'snap_tolerance': 3,
                    'join_tolerance': 3,
                    'edge_min_length': 3,
                    'min_words_vertical': 1,
                    'min_words_horizontal': 1,
                }
                
                # Essayer d'extraire les tableaux
                tables = page.extract_tables(table_settings)
                
                if tables:
                    for table in tables:
                        if table and len(table) > 1:
                            # Analyser l'en-tête
                            header = [str(cell).strip().lower() if cell else '' for cell in table[0]]
                            
                            # Trouver les index des colonnes
                            designation_idx = -1
                            dci_idx = -1
                            
                            for idx, col_name in enumerate(header):
                                if 'designation' in col_name or 'désignation' in col_name:
                                    designation_idx = idx
                                if 'dci' in col_name:
                                    dci_idx = idx
                            
                            # Extraire les données
                            if designation_idx >= 0 or dci_idx >= 0:
                                for row in table[1:]:
                                    if row and any(cell for cell in row):
                                        designation = row[designation_idx] if designation_idx >= 0 and designation_idx < len(row) else ''
                                        dci = row[dci_idx] if dci_idx >= 0 and dci_idx < len(row) else ''
                                        
                                        if designation or dci:
                                            all_data.append({
                                                'Page': page_num,
                                                'Designation': str(designation).strip() if designation else '',
                                                'DCI': str(dci).strip() if dci else ''
                                            })
                
                # Si pas de tableau avec lignes, essayer avec détection de texte
                else:
                    # Extraire le texte avec positions
                    words = page.extract_words(x_tolerance=3, y_tolerance=3)
                    
                    if not words:
                        continue
                    
                    # Grouper les mots par ligne (même y0)
                    lines_dict = {}
                    for word in words:
                        y_pos = round(word['top'])
                        if y_pos not in lines_dict:
                            lines_dict[y_pos] = []
                        lines_dict[y_pos].append(word)
                    
                    # Trier les lignes par position verticale
                    sorted_lines = sorted(lines_dict.items())
                    
                    # Chercher l'en-tête
                    header_y = None
                    designation_x = None
                    dci_x = None
                    
                    for y_pos, line_words in sorted_lines:
                        line_text = ' '.join([w['text'] for w in line_words]).lower()
                        
                        if 'designation' in line_text or 'désignation' in line_text:
                            header_y = y_pos
                            
                            # Trouver les positions X des colonnes
                            for w in line_words:
                                if 'designation' in w['text'].lower() or 'désignation' in w['text'].lower():
                                    designation_x = w['x0']
                                if 'dci' in w['text'].lower():
                                    dci_x = w['x0']
                            
                            break
                    
                    # Extraire les données si on a trouvé l'en-tête
                    if header_y and (designation_x or dci_x):
                        for y_pos, line_words in sorted_lines:
                            if y_pos <= header_y:
                                continue
                            
                            # Trier les mots de la ligne par position X
                            sorted_words = sorted(line_words, key=lambda w: w['x0'])
                            
                            # Séparer les mots selon les colonnes
                            designation_words = []
                            dci_words = []
                            
                            for word in sorted_words:
                                x_pos = word['x0']
                                
                                # Déterminer à quelle colonne appartient le mot
                                if designation_x and dci_x:
                                    # Calculer la distance aux deux colonnes
                                    dist_to_designation = abs(x_pos - designation_x)
                                    dist_to_dci = abs(x_pos - dci_x)
                                    
                                    if dist_to_designation < dist_to_dci and x_pos < dci_x:
                                        designation_words.append(word['text'])
                                    elif x_pos >= dci_x or dist_to_dci < dist_to_designation:
                                        dci_words.append(word['text'])
                                elif designation_x and x_pos < (designation_x + 200):
                                    designation_words.append(word['text'])
                                elif dci_x and x_pos >= dci_x:
                                    dci_words.append(word['text'])
                            
                            designation_text = ' '.join(designation_words).strip()
                            dci_text = ' '.join(dci_words).strip()
                            
                            # Ajouter si on a des données valides
                            if designation_text or dci_text:
                                # Éviter les en-têtes répétés
                                if 'designation' not in designation_text.lower() and 'page' not in designation_text.lower():
                                    all_data.append({
                                        'Page': page_num,
                                        'Designation': designation_text,
                                        'DCI': dci_text
                                    })
                
                if page_num % 50 == 0:
                    print(f"  📖 {page_num} pages traitées... ({len(all_data)} lignes extraites)")
    
    except Exception as e:
        print(f"❌ Erreur: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return all_data

def clean_data(data):
    """
    Nettoie les données extraites
    """
    cleaned = []
    
    for row in data:
        designation = row['Designation'].strip()
        dci = row['DCI'].strip()
        
        # Filtrer les lignes vides ou invalides
        if not designation and not dci:
            continue
        
        # Filtrer les en-têtes répétés
        if 'designation' in designation.lower() or 'nomenclature' in designation.lower():
            continue
        
        # Filtrer les numéros de page
        if re.match(r'^\d+$', designation):
            continue
        
        cleaned.append(row)
    
    return cleaned

def format_excel(excel_path):
    """
    Applique une mise en forme au fichier Excel
    """
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='Calibri', size=10)
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formater l'en-tête
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Formater les données
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # Ajuster les largeurs
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 70
        
        # Figer les volets
        ws.freeze_panes = ws['A2']
        
        wb.save(excel_path)
        print("✨ Mise en forme appliquée!")
        
    except Exception as e:
        print(f"⚠️ Erreur mise en forme: {e}")

def pdf_to_excel(pdf_path, excel_path):
    """
    Convertit le PDF en Excel
    """
    print("🚀 Extraction des données Designation et DCI...")
    print(f"📂 Fichier: {pdf_path}")
    print("-" * 60)
    
    # Extraction
    data = extract_table_data(pdf_path)
    
    if not data:
        print("❌ Aucune donnée trouvée!")
        return
    
    print(f"\n✓ {len(data)} lignes extraites brutes")
    
    # Nettoyage
    data = clean_data(data)
    print(f"✓ {len(data)} lignes après nettoyage")
    print("-" * 60)
    
    # Créer le DataFrame
    df = pd.DataFrame(data)
    
    # Sauvegarder
    df.to_excel(excel_path, index=False, engine='openpyxl')
    print(f"✅ Fichier créé: {excel_path}")
    
    # Mise en forme
    print("\n🎨 Application de la mise en forme...")
    format_excel(excel_path)
    
    print(f"\n📊 Résumé:")
    print(f"  - Total de lignes: {len(df)}")
    print(f"  - Pages avec données: {df['Page'].nunique()}")
    print(f"  - Designation non vides: {(df['Designation'] != '').sum()}")
    print(f"  - DCI non vides: {(df['DCI'] != '').sum()}")

if __name__ == "__main__":
    pdf_file = "NOMENCLATURE_NATIONALE_ANRP _ 2024.pdf"
    excel_file = "nomenclature_propre.xlsx"
    
    pdf_to_excel(pdf_file, excel_file)
